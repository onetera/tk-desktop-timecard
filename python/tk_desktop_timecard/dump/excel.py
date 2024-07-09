# -*- coding: utf-8 -*-

from .dump_constant import *

import openpyxl

from sgtk.platform.qt import QtGui, QtCore


HEADER_MAP = {
    'Id': 'check',
    'Person': 'user',
    'Link': 'task',
    'Link > Task > Link': 'shot/asset',
    'Date': 'date',
    'Duration': 'hour',
    'Project': 'project',
    'Description': 'description'
}


class ExcelLoad:
    def __init__(self, excel_file):
        self.row_infos = []
        self.excel_file = excel_file

        self.confirm_sync = False

        if self.header_confirm():
            self.read_excel()
            self.confirm_sync = True

    def read_excel(self):
        wb = openpyxl.load_workbook(self.excel_file)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = [cell for cell in row]
            if row_data[0]:
                check = True
            else:
                check = False
            row_data[0] = QtGui.QCheckBox()
            row_data[0].setChecked(check)

            self.row_infos.append(row_data)
        wb.close()
    
    def header_confirm(self):
        wb = openpyxl.load_workbook(self.excel_file)
        sheet = wb.active
        head_labels = [cell.value for cell in sheet[1]]


        for idx, label in enumerate(head_labels):
            if label in HEADER_MAP:
                head_labels[idx] = HEADER_MAP[label]

        if len(head_labels) != len(MODEL_KEYS.keys()):
            QtGui.QMessageBox.critical(None, "Error", u"excel file의 양식과 로드 된 파일의 행 갯수가 다릅니다.")
            wb.close()
            return False

        for idx, key in enumerate(MODEL_KEYS.keys()):
            if head_labels[idx].lower() != key.lower():
                QtGui.QMessageBox.critical(None, "Error", u"{0}번째 행의 '{1}'의 값이 '{2}'가 아닙니다.".format(idx+1 ,head_labels[idx], key))
                wb.close()
                return False
        
        wb.close()
        return True